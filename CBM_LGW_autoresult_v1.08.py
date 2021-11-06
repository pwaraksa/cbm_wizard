import csv
import datetime
import os
import pyodbc
import tempfile
import time
from tkinter import *
from tkinter import filedialog

#comment created by git pull

import openpyxl
import pandas as pd
import psutil
import xlsxwriter
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.chart import (LineChart, Reference, )
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font

# sprawdzenie zajecia pamieci
process = psutil.Process(os.getpid())
def mem(): print(f'{process.memory_info().rss:,}')
mem()

#pusta_lista_arodes - wypelniania pozniej
ls_modified_arodes = []

def GUI():
    global root, proces, button3, button4, arodes_nr, entry_1
    root = Tk()
    root.title("WIZARD CBM - LGW version")
    # root.geometry("500x500")

    # centering window
    w = 340
    h = 500
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))

    arodes_nr = ""

    # eng
    # button1 = Button(root, text="Select project file - additionality", command=SelectFile)
    # button1.pack(side="top", fill='both', expand=False, padx=20, pady=20)
    #
    # button5 = Button(root, text="Select project file - no additional effect", command=SelectFile2)
    # button5.pack(side="top", fill='both', expand=False, padx=20, pady=20)
    #
    # button5 = Button(root, text="Select input file", command=SelectInputFile)
    # button5.pack(side="top", fill='both', expand=False, padx=20, pady=20)
    #
    # button2 = Button(root, text="Select output folder", command=SelectOutputFolder)
    # button2.pack(side="top", fill='both', expand=False, padx=20, pady=20)

    # pl
    button1 = Button(root, text="Wybierz projekt - działania dodatkowe", command=SelectFile)
    button1.pack(side="top", fill='both', expand=False, padx=20, pady=20)

    button5 = Button(root, text="Wybierz projekt - bez działań dodatkowych", command=SelectFile2)
    button5.pack(side="top", fill='both', expand=False, padx=20, pady=20)

    button5 = Button(root, text="Wybierz plik wejściowy", command=SelectInputFile)
    button5.pack(side="top", fill='both', expand=False, padx=20, pady=20)

    button2 = Button(root, text="Wybierz folder do wygenerowania wyników", command=SelectOutputFolder)
    button2.pack(side="top", fill='both', expand=False, padx=20, pady=20)

    arodes_number = StringVar()

    label_1 = Label(root, text="Wpisz nr-y wewnętrzne wydzieleń do \n wygenerowania wykresów (opcjonalnie)")
    entry_1 = Entry(root, textvariable=arodes_number)
    label_1.pack()
    entry_1.pack()

    #eng
    # button4 = Button(root, text="Submit", command=submitArodes)
    # button4.pack()

    #pl
    button4 = Button(root, text="Wczytaj", command=submitArodes)
    button4.pack()

    #eng
    # button3 = Button(root, text="Generate results", command=startProcess)
    # button3.pack(side="top", fill='both', expand=False, padx=20, pady=20)

    #pl
    button3 = Button(root, text="Generuj wyniki", command=startProcess)
    button3.pack(side="top", fill='both', expand=False, padx=20, pady=20)

    arodes = Label(root, text="")
    arodes.pack()

    arodes_no = Label(root, text="")
    arodes_no.pack()

    proces = Label(root, text="")
    proces.pack()

    status = Label(root, text="ver. 1.08PL, author: Patryk Waraksa, p.waraksa@ibles.waw.pl", bd=1, relief=SUNKEN,
                   anchor=W)
    status.pack(side=BOTTOM, fill=X)

    ICON = (b'\x00\x00\x01\x00\x01\x00\x10\x10\x00\x00\x01\x00\x08\x00h\x05\x00\x00'
            b'\x16\x00\x00\x00(\x00\x00\x00\x10\x00\x00\x00 \x00\x00\x00\x01\x00'
            b'\x08\x00\x00\x00\x00\x00@\x05\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00'
            b'\x00\x01\x00\x00\x00\x01') + b'\x00' * 1282 + b'\xff' * 64

    _, ICON_PATH = tempfile.mkstemp()
    with open(ICON_PATH, 'wb') as icon_file:
        icon_file.write(ICON)

    root.iconbitmap(default=ICON_PATH)

    root.mainloop()


def startProcess():
    button3.configure(state=DISABLED)
    proces.configure(text="Processing... 0%")
    root.update()
    print("start: ")
    print(datetime.datetime.now())

    # obliczenie powierzchni z pliku wsadowego i lista arodes
    Process0()

    # zaladowanie SQL
    Agregacje_SQL()

    #english
    # # agregacja po time_step
    # Process1(filename1, "tblPoolIndicators_groupby_TimeStep_dz_dod.xlsx")
    # Process1(filename2, "tblPoolIndicators_groupby_TimeStep_bez_dod.xlsx")
    #
    # # agregacja po nadlesnictwie
    # Process2(filename1, "tblPoolIndicators_groupby_KodNadl_dz_dod.xlsx")
    # Process2(filename2, "tblPoolIndicators_groupby_KodNadl_bez_dod.xlsx")
    # Process2_ha(filename1, "tblPoolIndicators_groupby_KodNadl_dz_dod_ha.xlsx")
    # Process2_ha(filename2, "tblPoolIndicators_groupby_KodNadl_bez_dod_ha.xlsx")
    # Process2_roznica("tblPoolIndicators_groupby_KodNadl_roznica.xlsx")
    # Process2_roznica_ha("tblPoolIndicators_groupby_KodNadl_roznica_ha.xlsx")
    #
    # # agregacja po dzialaniu dodatkowym
    # Process3(filename1, "tblPoolIndicators_groupby_DzDod_dz_dod.xlsx")
    # Process3(filename2, "tblPoolIndicators_groupby_DzDod_bez_dod.xlsx")
    # Process3_ha(filename1, "tblPoolIndicators_groupby_DzDod_dz_dod_ha.xlsx")
    # Process3_ha(filename2, "tblPoolIndicators_groupby_DzDod_bez_dod_ha.xlsx")
    # Process3_roznica("tblPoolIndicators_groupby_DzDod_roznica.xlsx")
    # Process3_roznica_ha("tblPoolIndicators_groupby_DzDod_roznica_ha.xlsx")
    #
    # # agregacja po arodes - tylko jesli zdefiniowana jest lista arodes
    # # if ls_modified_arodes != []:
    # Process4(filename1, "tblPoolIndicators_groupby_Arodes_dz_dod.xlsx")
    # Process4(filename2, "tblPoolIndicators_groupby_Arodes_bez_dod.xlsx")
    # Process4_ha(filename1, "tblPoolIndicators_groupby_Arodes_dz_dod_ha.xlsx")
    # Process4_ha(filename2, "tblPoolIndicators_groupby_Arodes_bez_dod_ha.xlsx")
    # Process4_roznica("tblPoolIndicators_groupby_Arodes_roznica.xlsx")
    # Process4_roznica_ha("tblPoolIndicators_groupby_Arodes_roznica_ha.xlsx")
    #
    # # bez agregacji - wszystkie dane
    # Process5(filename1, "tblPoolIndicators_dz_dod.xlsx")
    #
    # Process5(filename2, "tblPoolIndicators_bez_dod.xlsx")

    #pl
    # agregacja po time_step
    Process1(filename1, "Zasoby_wegla_agreg_RokZaburz_dz_dod.xlsx")
    Process1(filename2, "Zasoby_wegla_agreg_RokZaburz_bez_dod.xlsx")
    #
    # # agregacja po nadlesnictwie
    Process2(filename1, "Zasoby_wegla_agreg_KodNadl_dz_dod.xlsx")
    Process2(filename2, "Zasoby_wegla_agreg_KodNadl_bez_dod.xlsx")
    Process2_ha(filename1, "Zasoby_wegla_agreg_KodNadl_dz_dod_ha.xlsx")
    Process2_ha(filename2, "Zasoby_wegla_agreg_KodNadl_bez_dod_ha.xlsx")
    Process2_roznica("Zasoby_wegla_agreg_KodNadl_roznica.xlsx")
    Process2_roznica_ha("Zasoby_wegla_agreg_KodNadl_roznica_ha.xlsx")

    # agregacja po dzialaniu dodatkowym
    Process3(filename1, "Zasoby_wegla_agreg_DzDod_dz_dod.xlsx")
    Process3(filename2, "Zasoby_wegla_agreg_DzDod_bez_dod.xlsx")
    Process3_ha(filename1, "Zasoby_wegla_agreg_DzDod_dz_dod_ha.xlsx")
    Process3_ha(filename2, "Zasoby_wegla_agreg_DzDod_bez_dod_ha.xlsx")
    Process3_roznica("Zasoby_wegla_agreg_DzDod_roznica.xlsx")
    Process3_roznica_ha("Zasoby_wegla_agreg_DzDod_roznica_ha.xlsx")

    # agregacja po arodes - tylko jesli zdefiniowana jest lista arodes
    # if ls_modified_arodes != []:
    # Process4(filename1, "Zasoby_wegla_agreg_Arodes_dz_dod.xlsx")
    # Process4(filename2, "Zasoby_wegla_agreg_Arodes_bez_dod.xlsx")
    # Process4_ha(filename1, "Zasoby_wegla_agreg_Arodes_dz_dod_ha.xlsx")
    # Process4_ha(filename2, "Zasoby_wegla_agreg_Arodes_bez_dod_ha.xlsx")
    # Process4_roznica("Zasoby_wegla_agreg_Arodes_roznica.xlsx")
    # Process4_roznica_ha("Zasoby_wegla_agreg_Arodes_roznica_ha.xlsx")

    # bez agregacji - wszystkie dane
    Process5(filename1, "Zasoby_wegla_dz_dod.xlsx")
    Process5(filename2, "Zasoby_wegla_bez_dod.xlsx")


    # komunikat o zakończeniu procesu, zwolnienie pamięci
    Process6()

    # proba symultanicznych procesow
    # if __name__ == '__main__':
    #
    #     from multiprocessing import Process
    #
    #     p1= Process(target=Process1)
    #     p2 = Process(target=Process2)
    #     p3 = Process(target=Process3)
    #     p1.start()
    #     p2.start()
    #     p3.start()
    #     p1.join()
    #     p2.join()
    #     p3.join()


def Process0():
    # dane z pliku wsadowego, obliczenie zagregowanej powierzchni, zdefiniowanie listy wydzielen do wynikow

    global DzDod_pow, Nadl_pow, ls_modified_arodes

    # lista arodes do wygenerowania wynikow w Proces4 - agregacja po arodes

    # ls_modified_arodes = [1606021096,1606028471,1606028477,1606021356,1420024418,1420031311,1420007759,1420008710,1420009140,1420009271,1420024506,1420024257,1420031425,1420031426,1420007991,1420031319,1407019066,1407009709,1407018476,1407018478,1407021017,1407041428,1407041766,1318047303,1318003276,1318005675,1318010056,1318010491,1318012529,1204001875,1204054980,1008017172,1008040646,1008043414,1008043494,1008043329,1008027707,1008043379,1008043427,915023186,915023020,915022994,915023723,817054146,817054147,817054149,817054152,817054153,817015090,817054359,817054360,817054361,817062909,817062985,817062871,817062935,817053378,816001914,816001924,816001926,816001931,816001938,816002046,816002048,816002070,816002074,816002076,816003190,816003201,816003293,816009301,816004802,816004812,816018437,816004945,816004950,816018451,816005068,816005071,816005079,816005081,816005083,816018456,108029988,108018369,121022398,121022585,216023593,216023516,216023378,216023534,216023100,235000547,235013762,237000778,237001877,237012453,237002922,811003752,811004011,811042483,811042485,811042962,811041812,811004196,811004205,811003908,811041866,811041867,811041877,811004068,811041883,811004078,811004080,811041886,811004218,811041900,811008415,811008428,811008422,811008424,811008436,811008769,811008777,811042461,811008950,811008961,811042939,811039881,1701007998,1701016287,1701018089,1701009006,1701009092,1701009154,1701010383,1701009928,1701009964,1701010777,1701018613,1701012731,1701013204,1701008210,1420031189,1420024610,1420026832,1420025665,1420025666,1420024921,1420000967,1420001724,1420000788,1420019434,1420024294,1420003138,1420018118,1420003198,1420003818,1420002562,1420004329,1420004499,1420024460,1420024532,1420004818,1420029936,1420029958,1420031305,1420019264,1420029873,1420019267,1420029875,1420020579,1420029955,1420024235,1420006036,1420006155,1420021485,1420024566,1420024640,1420026338,1420024416,1420027145,1420026281,1420008621,1420008699,1420008796,1420008020,1420025162,1420021366,1420008065,1420008888,1420008241,1420008251,1420009206,1420009292,1420009672,1420010172,1420022430,1420010436,1420010669,1420008000,1420028009,1420011486,1420024456,1420024358,1420028033,1420019631,1420004092,1420021037,1413042119,1413005113,1413005814,1413067130,1204055875,1204028757,1204034962,1204056517,1204056666,1204054897,1204036078,1204009760,1204057075,1204027070,1204011181,1204055559,1204054542,1204005372,1204011950,1204054751,1204006710,1008043728,1008043712,1008043762,1008043764,915021347,120035643,120035749,120014705,120037463,120005897,120008988,120009228,120009547,120011323,120013038,120012668,120036519,120014348,120014469,120014558,206002611,206018304,216023070,216023076,235013485,1701007987,1701008004,1701008098,1701008108,1701016942,1701008143,1701008202,1701008204,1701008254,1701008261,1701008285,1701018697,1701018088,1701008886,1701009029,1701009042,1701009062,1701009103,1701009105,1701009107,1701009328,1701009404,1701009406,1701009439,1701009441,1701009450,1701009474,1701009484,1701009557,1701010340,1701009730,1701017600,1701009796,1701010029,1701010433,1701010435,1701010726,1701010728,1701010945,1701010969,1701010975,1701017060,1701011677,1701012238,1701012345,1701018632,1701012377,1701012379,1701012381,1701012389,1701012450,1701012469,1701012716,1701012746,1701012855,1701012870,1701012875,1701017778,1701013125,1701013272,1701013379,1701015824,1407011447,1407011461,1420000122,1008027294,817060808,817060810,817062908,817061124,817015536,817054836,817054843,817015121,817054351,817054352,817060836,817016631,817053981,817053982,817054151,817053392,817060911,817053415,206005272,216023090,235001339,235013556,1407005267,1204028294,1204054527,1420018885,1420000913,1204001241,1204034250,1204034937,1204029081,1204027601,1204035376,1204056810,1204054566,1204011766,1204011776,1204035881,1204028149,1204035602,1204057314,1204009142,1204034201,1204001239,1204035368,1204011822,1204035861,1204008758,1204035650,1204035651,1008021810,1008017196,1008043500,1008043607,1008043449,1008043589,1008041364,1008040050,1008023674,1008043609,1008043583,1008043471,108011984,206015255,206000813,206000855,206003263,206005401,206009116,206009338,216008462,235013554]

    proces.configure(text="Processing... 5%")
    root.update()

    df = pd.read_excel(io=input_filename, sheet_name="SIT_Inventory")

    dz_dod_list = list(df.dzial_dod)
    kod_nadl_list = list(df.kod_nadl)
    # arodes_list = list(df.arodes)
    area_list = list(df.Area)

    list_of_tuples = list(zip(dz_dod_list, kod_nadl_list, area_list))
    # print(list_of_tuples)

    df = pd.DataFrame(list_of_tuples, columns=['DzDod', 'Nadl', 'Area'])

    sum_by_group = df.groupby(['DzDod']).sum()
    # print(sum_by_group)
    dict = sum_by_group.to_dict('dict')
    # print(dict)
    dict_inside = dict['Area']
    # print(dict_inside)
    DzDod_pow = dict_inside
    print(DzDod_pow)

    sum_by_group = df.groupby(['Nadl']).sum()
    dict = sum_by_group.to_dict('dict')
    dict_inside = dict['Area']
    Nadl_pow = dict_inside
    print(Nadl_pow)

    # sum_by_group = df.groupby(['Arodes']).sum()
    # dict = sum_by_group.to_dict('dict')
    # dict_inside = dict['Area']
    # Arodes_pow = dict_inside
    # # print(Arodes_pow)

    print(f"The areas from input file are calculated. Time: {datetime.datetime.now()}")

    proces.configure(text="Processing... 7%")
    root.update()


def Agregacje_SQL():
    global SQL_agreg_time_step, SQL_agreg_kod_nadl, SQL_agreg_dzial_dod, SQL_agreg_arodes, SQL_tblPoolIndicators

    SQL_agreg_time_step = '''SELECT TimeStep AS Rok_symulacji,

        Sum_VFastAG+ 
        Sum_VFastBG+ 
        Sum_FastAG+ 
        Sum_FastBG+ 
        Sum_Medium+ 
        Sum_SlowAG+ 
        Sum_SlowBG+ 
        Sum_SWStemSnag+ 
        Sum_SWBranchSnag+ 
        Sum_HWStemSnag+ 
        Sum_HWBranchSnag+ 
        Sum_SW_Merch+ 
        Sum_SW_Foliage+ 
        Sum_SW_Other+ 
        Sum_SW_Coarse+ 
        Sum_SW_Fine+ 
        Sum_HW_Merch+ 
        Sum_HW_Foliage+ 
        Sum_HW_Other+ 
        Sum_HW_Coarse+ 
        Sum_HW_Fine AS Całkowity_ekosystem,

        Sum_SW_Merch+ 
        Sum_SW_Foliage+ 
        Sum_SW_Other+ 
        Sum_HW_Merch+ 
        Sum_HW_Foliage+ 
        Sum_HW_Other+
        Sum_SW_Coarse+ 
        Sum_SW_Fine+ 
        Sum_HW_Coarse+ 
        Sum_HW_Fine AS Całkowita_biomasa,

        Sum_SW_Merch+ 
        Sum_HW_Merch+ 
        Sum_HW_Fine AS Grubizna_iglasta_i_liściasta,

        Sum_SW_Merch+ 
        Sum_SW_Foliage+ 
        Sum_SW_Other+ 
        Sum_HW_Merch+ 
        Sum_HW_Foliage+ 
        Sum_HW_Other AS Biomasa_nadziemna,

        Sum_SW_Coarse+ 
        Sum_SW_Fine+ 
        Sum_HW_Coarse+ 
        Sum_HW_Fine AS Biomasa_podziemna,

        Sum_VFastAG+ 
        Sum_FastAG+ 
        Sum_Medium+ 
        Sum_SlowAG+ 
        Sum_SWStemSnag+ 
        Sum_SWBranchSnag+ 
        Sum_HWStemSnag+ 
        Sum_HWBranchSnag+
        Sum_VFastBG+ 
        Sum_FastBG+ 
        Sum_SlowBG AS Martwa_materia_organiczna_DOM,

        Sum_VFastAG+ 
        Sum_FastAG+ 
        Sum_Medium+ 
        Sum_SlowAG+ 
        Sum_SWStemSnag+ 
        Sum_SWBranchSnag+ 
        Sum_HWStemSnag+ 
        Sum_HWBranchSnag AS Nadziemna_martwa_mat_org_DOM,

        Sum_VFastBG+ 
        Sum_FastBG+ 
        Sum_SlowBG AS Podziemna_martwa_mat_org_DOM,

        Sum_FastBG+ 
        Sum_Medium+ 
        Sum_SWStemSnag+ 
        Sum_SWBranchSnag+ 
        Sum_HWStemSnag+ 
        Sum_HWBranchSnag AS Martwe_drewno,

        Sum_VFastAG+ 
        Sum_FastAG+ 
        Sum_SlowAG AS Ścioła,

        Sum_VFastBG+ 
        Sum_SlowBG AS Węgiel_w_glebie,

        Sum_VFastAG AS Nadziemna_b_sz_r_m_m_o_DOM,
        Sum_VFastBG AS Podziemna_b_sz_r_m_m_o_DOM,
        Sum_FastAG AS Nadziemna_sz_r_m_m_o_DOM,
        Sum_FastBG AS Podziemna_sz_r_m_m_o_DOM,
        Sum_Medium AS Średnio_r_m_m_o_DOM,
        Sum_SlowAG AS Nadziemna_wolno_r_m_m_o_DOM,
        Sum_SlowBG AS Podziemna_wolno_r_m_m_o_DOM, 
        Sum_SWStemSnag AS Iglasty_posusz_pnie,
        Sum_SWBranchSnag AS Iglasty_posusz_gałęzie,
        Sum_HWStemSnag AS Liściasty_posusz_pnie,
        Sum_HWBranchSnag AS Liściasty_posusz_gałęzie,
        Sum_SW_Merch AS Grubizna_iglasta,
        Sum_SW_Foliage AS Aparat_asymilacyjny_igl,
        Sum_SW_Other AS Inne_iglaste,
        Sum_SW_Coarse AS Grube_korzenie_iglaste,
        Sum_SW_Fine AS Cienkie_korzenie_iglaste,
        Sum_HW_Merch AS Grubizna_liściasta,
        Sum_HW_Foliage AS Aparat_asymilacyjny_liśc,
        Sum_HW_Other AS Inne_liściaste,
        Sum_HW_Coarse AS Grube_korzenie_liściaste,
        Sum_HW_Fine AS Cienkie_korzenie_liściaste_HW

        FROM
        (SELECT tblPoolIndicators.TimeStep,
        Sum(tblPoolIndicators.VFastAG) AS Sum_VFastAG,
        Sum(tblPoolIndicators.VFastBG) AS Sum_VFastBG,
        Sum(tblPoolIndicators.FastAG) AS Sum_FastAG,
        Sum(tblPoolIndicators.FastBG) AS Sum_FastBG,
        Sum(tblPoolIndicators.Medium) AS Sum_Medium,
        Sum(tblPoolIndicators.SlowAG) AS Sum_SlowAG,
        Sum(tblPoolIndicators.SlowBG) AS Sum_SlowBG,
        Sum(tblPoolIndicators.SWStemSnag) AS Sum_SWStemSnag,
        Sum(tblPoolIndicators.SWBranchSnag) AS Sum_SWBranchSnag,
        Sum(tblPoolIndicators.HWStemSnag) AS Sum_HWStemSnag,
        Sum(tblPoolIndicators.HWBranchSnag) AS Sum_HWBranchSnag,
        Sum(tblPoolIndicators.SW_Merch) AS Sum_SW_Merch,
        Sum(tblPoolIndicators.SW_Foliage) AS Sum_SW_Foliage,
        Sum(tblPoolIndicators.SW_Other) AS Sum_SW_Other,
        Sum(tblPoolIndicators.SW_Coarse) AS Sum_SW_Coarse,
        Sum(tblPoolIndicators.SW_Fine) AS Sum_SW_Fine,
        Sum(tblPoolIndicators.HW_Merch) AS Sum_HW_Merch,
        Sum(tblPoolIndicators.HW_Foliage) AS Sum_HW_Foliage,
        Sum(tblPoolIndicators.HW_Other) AS Sum_HW_Other,
        Sum(tblPoolIndicators.HW_Coarse) AS Sum_HW_Coarse,
        Sum(tblPoolIndicators.HW_Fine) AS Sum_HW_Fine
        FROM tblUserDefdClassSets INNER JOIN tblPoolIndicators
        ON tblUserDefdClassSets.UserDefdClassSetID = tblPoolIndicators.UserDefdClassSetID
        WHERE (((tblUserDefdClassSets.Name)<>'SO,999999,99-99,REF,9999999999'))
        GROUP BY tblPoolIndicators.TimeStep)'''

    SQL_agreg_kod_nadl = '''SELECT  KodNadl, TimeStep AS Rok_symulacji, 
        Sum(Total_Ecosystem) As Całkowity_ekosystem, 
        Sum(Total_Biomass) As Całkowita_biomasa, 
        Sum(Total_Merch) AS Grubizna_iglasta_i_liściasta,
        Sum(Aboveground_Biomass) As Biomasa_nadziemna, 
        Sum(Belowground_Biomass) AS Biomasa_podziemna, 
        Sum(Total_DOM) AS Martwa_materia_organiczna_DOM, 
    	Sum(Aboveground_DOM) AS Nadziemna_martwa_mat_org_DOM,
        Sum(Belowground_DOM) AS Podziemna_martwa_mat_org_DOM, 
        Sum(Deadwood) AS Martwe_drewno, 
        Sum(Litter) AS Ścioła, 
        Sum(Soil_C) AS Węgiel_w_glebie,

    	Sum(tblPoolIndicators.VFastAG) AS Nadziemna_b_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.VFastBG) AS Podziemna_b_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.FastAG) AS Nadziemna_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.FastBG) AS Podziemna_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.Medium) AS Średnio_r_m_m_o_DOM,
        Sum(tblPoolIndicators.SlowAG) AS Nadziemna_wolno_r_m_m_o_DOM,
        Sum(tblPoolIndicators.SlowBG) AS Podziemna_wolno_r_m_m_o_DOM, 
        Sum(tblPoolIndicators.SWStemSnag) AS Iglasty_posusz_pnie,
        Sum(tblPoolIndicators.SWBranchSnag) AS Iglasty_posusz_gałęzie,
        Sum(tblPoolIndicators.HWStemSnag) AS Liściasty_posusz_pnie,
        Sum(tblPoolIndicators.HWBranchSnag) AS Liściasty_posusz_gałęzie,
        Sum(tblPoolIndicators.SW_Merch) AS Grubizna_iglasta,
        Sum(tblPoolIndicators.SW_Foliage) AS Aparat_asymilacyjny_igl,
        Sum(tblPoolIndicators.SW_Other) AS Inne_iglaste,
        Sum(tblPoolIndicators.SW_Coarse) AS Grube_korzenie_iglaste,
        Sum(tblPoolIndicators.SW_Fine) AS Cienkie_korzenie_iglaste,
        Sum(tblPoolIndicators.HW_Merch) AS Grubizna_liściasta,
        Sum(tblPoolIndicators.HW_Foliage) AS Aparat_asymilacyjny_liśc,
        Sum(tblPoolIndicators.HW_Other) AS Inne_liściaste,
        Sum(tblPoolIndicators.HW_Coarse) AS Grube_korzenie_liściaste,
        Sum(tblPoolIndicators.HW_Fine) AS Cienkie_korzenie_liściaste




        FROM

        (SELECT tblUserDefdClassSets.Name,

        Left([Name],InStr([Name],',')-1) AS Species, 
        Right((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)),Len((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)))-InStr((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)),',')) AS IdGatWydz, 
        Right(Left([Name],InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')-1),5) AS KodNadl, 
        Right(Left([Name],InStr(InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')+1,[Name],',')-1),Len(Left([Name],InStr(InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')+1,[Name],',')-1))-InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')) AS DzDod, 
        Right([Name],Len([Name])-InStr(InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')+1,[Name],',')) AS Arodes,

        tblPoolIndicators.UserDefdClassSetID, 
        tblPoolIndicators.PoolIndID, 
        tblPoolIndicators.TimeStep,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SlowBG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag+ 
        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Total_Ecosystem,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other+
        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Total_Biomass,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Fine AS Total_Merch,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other AS Aboveground_Biomass,

        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Belowground_Biomass,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag+
        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.SlowBG AS Total_DOM,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag AS Aboveground_DOM,

        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.SlowBG AS Belowground_DOM,

        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag AS Deadwood,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.SlowAG AS Litter,

        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.SlowBG AS Soil_C,

        tblPoolIndicators.VFastAG, 
        tblPoolIndicators.VFastBG, 
        tblPoolIndicators.FastAG, 
        tblPoolIndicators.FastBG, 
        tblPoolIndicators.Medium, 
        tblPoolIndicators.SlowAG, 
        tblPoolIndicators.SlowBG, 
        tblPoolIndicators.SWStemSnag, 
        tblPoolIndicators.SWBranchSnag, 
        tblPoolIndicators.HWStemSnag, 
        tblPoolIndicators.HWBranchSnag, 
        tblPoolIndicators.SW_Merch, 
        tblPoolIndicators.SW_Foliage, 
        tblPoolIndicators.SW_Other, 
        tblPoolIndicators.SW_Coarse, 
        tblPoolIndicators.SW_Fine, 
        tblPoolIndicators.HW_Merch, 
        tblPoolIndicators.HW_Foliage, 
        tblPoolIndicators.HW_Other, 
        tblPoolIndicators.HW_Coarse, 
        tblPoolIndicators.HW_Fine

        FROM tblUserDefdClassSets INNER JOIN tblPoolIndicators 
        ON tblUserDefdClassSets.UserDefdClassSetID = tblPoolIndicators.UserDefdClassSetID
        WHERE (((tblUserDefdClassSets.Name)<>'SO,999999,99-99,REF,9999999999'))
        ORDER BY tblPoolIndicators.PoolIndID, tblPoolIndicators.TimeStep)

        GROUP BY KodNadl, Timestep'''

    SQL_agreg_dzial_dod = '''SELECT DzDod, TimeStep As Rok_symulacji, 
        Sum(Total_Ecosystem) As Całkowity_ekosystem, 
        Sum(Total_Biomass) As Całkowita_biomasa, 
        Sum(Total_Merch) AS Grubizna_iglasta_i_liściasta, 
        Sum(Aboveground_Biomass) As Biomasa_nadziemna, 
        Sum(Belowground_Biomass) AS Biomasa_podziemna, 
        Sum(Total_DOM) AS Martwa_materia_organiczna_DOM, 
        Sum(Aboveground_DOM) AS Nadziemna_martwa_mat_org_DOM,
        Sum(Belowground_DOM) AS Podziemna_martwa_mat_org_DOM, 
        Sum(Deadwood) AS Martwe_drewno, 
        Sum(Litter) AS Ścioła, 
        Sum(Soil_C) AS Węgiel_w_glebie,

        Sum(tblPoolIndicators.VFastAG) AS Nadziemna_b_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.VFastBG) AS Podziemna_b_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.FastAG) AS Nadziemna_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.FastBG) AS Podziemna_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.Medium) AS Średnio_r_m_m_o_DOM,
        Sum(tblPoolIndicators.SlowAG) AS Nadziemna_wolno_r_m_m_o_DOM,
        Sum(tblPoolIndicators.SlowBG) AS Podziemna_wolno_r_m_m_o_DOM, 
        Sum(tblPoolIndicators.SWStemSnag) AS Iglasty_posusz_pnie,
        Sum(tblPoolIndicators.SWBranchSnag) AS Iglasty_posusz_gałęzie,
        Sum(tblPoolIndicators.HWStemSnag) AS Liściasty_posusz_pnie,
        Sum(tblPoolIndicators.HWBranchSnag) AS Liściasty_posusz_gałęzie,
        Sum(tblPoolIndicators.SW_Merch) AS Grubizna_iglasta,
        Sum(tblPoolIndicators.SW_Foliage) AS Aparat_asymilacyjny_igl,
        Sum(tblPoolIndicators.SW_Other) AS Inne_iglaste,
        Sum(tblPoolIndicators.SW_Coarse) AS Grube_korzenie_iglaste,
        Sum(tblPoolIndicators.SW_Fine) AS Cienkie_korzenie_iglaste,
        Sum(tblPoolIndicators.HW_Merch) AS Grubizna_liściasta,
        Sum(tblPoolIndicators.HW_Foliage) AS Aparat_asymilacyjny_liśc,
        Sum(tblPoolIndicators.HW_Other) AS Inne_liściaste,
        Sum(tblPoolIndicators.HW_Coarse) AS Grube_korzenie_liściaste,
        Sum(tblPoolIndicators.HW_Fine) AS Cienkie_korzenie_liściaste




        FROM

        (SELECT tblUserDefdClassSets.Name,

        Left([Name],InStr([Name],',')-1) AS Species, 
        Right((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)),Len((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)))-InStr((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)),',')) AS IdGatWydz, 
        Right(Left([Name],InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')-1),5) AS KodNadl, 
        Right([Name],Len([Name])-InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')) AS DzDod,

        tblPoolIndicators.UserDefdClassSetID, 
        tblPoolIndicators.PoolIndID, 
        tblPoolIndicators.TimeStep,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SlowBG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag+ 
        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Total_Ecosystem,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other+
        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Total_Biomass,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Fine AS Total_Merch,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other AS Aboveground_Biomass,

        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Belowground_Biomass,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag+
        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.SlowBG AS Total_DOM,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag AS Aboveground_DOM,

        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.SlowBG AS Belowground_DOM,

        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag AS Deadwood,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.SlowAG AS Litter,

        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.SlowBG AS Soil_C,

        tblPoolIndicators.VFastAG, 
        tblPoolIndicators.VFastBG, 
        tblPoolIndicators.FastAG, 
        tblPoolIndicators.FastBG, 
        tblPoolIndicators.Medium, 
        tblPoolIndicators.SlowAG, 
        tblPoolIndicators.SlowBG, 
        tblPoolIndicators.SWStemSnag, 
        tblPoolIndicators.SWBranchSnag, 
        tblPoolIndicators.HWStemSnag, 
        tblPoolIndicators.HWBranchSnag, 
        tblPoolIndicators.SW_Merch, 
        tblPoolIndicators.SW_Foliage, 
        tblPoolIndicators.SW_Other, 
        tblPoolIndicators.SW_Coarse, 
        tblPoolIndicators.SW_Fine, 
        tblPoolIndicators.HW_Merch, 
        tblPoolIndicators.HW_Foliage, 
        tblPoolIndicators.HW_Other, 
        tblPoolIndicators.HW_Coarse, 
        tblPoolIndicators.HW_Fine

        FROM tblUserDefdClassSets INNER JOIN tblPoolIndicators 
        ON tblUserDefdClassSets.UserDefdClassSetID = tblPoolIndicators.UserDefdClassSetID
        WHERE (((tblUserDefdClassSets.Name)<>'SO,999999,99-99,REF,9999999999'))
        ORDER BY tblPoolIndicators.PoolIndID, tblPoolIndicators.TimeStep)

        GROUP BY DzDod, Timestep'''

    SQL_agreg_arodes = '''SELECT Arodes, TimeStep AS Rok_symulacji, 
        Sum(Total_Ecosystem) As Całkowity_ekosystem, 
        Sum(Total_Biomass) As Całkowita_biomasa, 
        Sum(Total_Merch) AS Grubizna_iglasta_i_liściasta, 
        Sum(Aboveground_Biomass) As Biomasa_nadziemna, 
        Sum(Belowground_Biomass) AS Biomasa_podziemna, 
        Sum(Total_DOM) AS Martwa_materia_organiczna_DOM, 
        Sum(Aboveground_DOM) AS Nadziemna_martwa_mat_org_DOM,
        Sum(Belowground_DOM) AS Podziemna_martwa_mat_org_DOM, 
        Sum(Deadwood) AS Martwe_drewno, 
        Sum(Litter) AS Ścioła, 
        Sum(Soil_C) AS Węgiel_w_glebie,

        Sum(tblPoolIndicators.VFastAG) AS Nadziemna_b_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.VFastBG) AS Podziemna_b_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.FastAG) AS Nadziemna_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.FastBG) AS Podziemna_sz_r_m_m_o_DOM,
        Sum(tblPoolIndicators.Medium) AS Średnio_r_m_m_o_DOM,
        Sum(tblPoolIndicators.SlowAG) AS Nadziemna_wolno_r_m_m_o_DOM,
        Sum(tblPoolIndicators.SlowBG) AS Podziemna_wolno_r_m_m_o_DOM, 
        Sum(tblPoolIndicators.SWStemSnag) AS Iglasty_posusz_pnie,
        Sum(tblPoolIndicators.SWBranchSnag) AS Iglasty_posusz_gałęzie,
        Sum(tblPoolIndicators.HWStemSnag) AS Liściasty_posusz_pnie,
        Sum(tblPoolIndicators.HWBranchSnag) AS Liściasty_posusz_gałęzie,
        Sum(tblPoolIndicators.SW_Merch) AS Grubizna_iglasta,
        Sum(tblPoolIndicators.SW_Foliage) AS Aparat_asymilacyjny_igl,
        Sum(tblPoolIndicators.SW_Other) AS Inne_iglaste,
        Sum(tblPoolIndicators.SW_Coarse) AS Grube_korzenie_iglaste,
        Sum(tblPoolIndicators.SW_Fine) AS Cienkie_korzenie_iglaste,
        Sum(tblPoolIndicators.HW_Merch) AS Grubizna_liściasta,
        Sum(tblPoolIndicators.HW_Foliage) AS Aparat_asymilacyjny_liśc,
        Sum(tblPoolIndicators.HW_Other) AS Inne_liściaste,
        Sum(tblPoolIndicators.HW_Coarse) AS Grube_korzenie_liściaste,
        Sum(tblPoolIndicators.HW_Fine) AS Cienkie_korzenie_liściaste_HW




        FROM

        (SELECT tblUserDefdClassSets.Name,

        Left([Name],InStr([Name],',')-1) AS Species, 
        Right((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)),Len((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)))-InStr((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)),',')) AS IdGatWydz, 
        Right(Left([Name],InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')-1),5) AS KodNadl, 
        Right(Left([Name],InStr(InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')+1,[Name],',')-1),Len(Left([Name],InStr(InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')+1,[Name],',')-1))-InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')) AS DzDod, 
        Right([Name],Len([Name])-InStr(InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')+1,[Name],',')) AS Arodes,

        tblPoolIndicators.UserDefdClassSetID, 
        tblPoolIndicators.PoolIndID, 
        tblPoolIndicators.TimeStep,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SlowBG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag+ 
        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Total_Ecosystem,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other+
        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Total_Biomass,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Fine AS Total_Merch,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other AS Aboveground_Biomass,

        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Belowground_Biomass,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag+
        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.SlowBG AS Total_DOM,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag AS Aboveground_DOM,

        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.SlowBG AS Belowground_DOM,

        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag AS Deadwood,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.SlowAG AS Litter,

        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.SlowBG AS Soil_C,

        tblPoolIndicators.VFastAG, 
        tblPoolIndicators.VFastBG, 
        tblPoolIndicators.FastAG, 
        tblPoolIndicators.FastBG, 
        tblPoolIndicators.Medium, 
        tblPoolIndicators.SlowAG, 
        tblPoolIndicators.SlowBG, 
        tblPoolIndicators.SWStemSnag, 
        tblPoolIndicators.SWBranchSnag, 
        tblPoolIndicators.HWStemSnag, 
        tblPoolIndicators.HWBranchSnag, 
        tblPoolIndicators.SW_Merch, 
        tblPoolIndicators.SW_Foliage, 
        tblPoolIndicators.SW_Other, 
        tblPoolIndicators.SW_Coarse, 
        tblPoolIndicators.SW_Fine, 
        tblPoolIndicators.HW_Merch, 
        tblPoolIndicators.HW_Foliage, 
        tblPoolIndicators.HW_Other, 
        tblPoolIndicators.HW_Coarse, 
        tblPoolIndicators.HW_Fine

        FROM tblUserDefdClassSets INNER JOIN tblPoolIndicators 
        ON tblUserDefdClassSets.UserDefdClassSetID = tblPoolIndicators.UserDefdClassSetID
        WHERE (((tblUserDefdClassSets.Name)<>'SO,999999,99-99,REF,9999999999'))
        ORDER BY tblPoolIndicators.PoolIndID, tblPoolIndicators.TimeStep)

        GROUP BY Arodes, Timestep'''

    SQL_tblPoolIndicators = '''SELECT tblUserDefdClassSets.Name AS Nazwa,

        Left([Name],InStr([Name],',')-1) AS Species, 
        Right((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)),Len((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)))-InStr((Left([Name],InStr(InStr([Name],',')+1,[Name],',')-1)),',')) AS IdGatWydz, 
        Right(Left([Name],InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')-1),5) AS KodNadl, 
        Right(Left([Name],InStr(InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')+1,[Name],',')-1),Len(Left([Name],InStr(InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')+1,[Name],',')-1))-InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')) AS DzDod, 
        Right([Name],Len([Name])-InStr(InStr(InStr(InStr([Name],',')+1,[Name],',')+1,[Name],',')+1,[Name],',')) AS Arodes,

        tblPoolIndicators.UserDefdClassSetID, 
        tblPoolIndicators.PoolIndID, 
        tblPoolIndicators.TimeStep,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SlowBG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag+ 
        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Całkowity_ekosystem,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other+
        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Całkowita_biomasa,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Fine AS Grubizna_iglasta_i_liściasta,

        tblPoolIndicators.SW_Merch+ 
        tblPoolIndicators.SW_Foliage+ 
        tblPoolIndicators.SW_Other+ 
        tblPoolIndicators.HW_Merch+ 
        tblPoolIndicators.HW_Foliage+ 
        tblPoolIndicators.HW_Other AS Biomasa_nadziemna,

        tblPoolIndicators.SW_Coarse+ 
        tblPoolIndicators.SW_Fine+ 
        tblPoolIndicators.HW_Coarse+ 
        tblPoolIndicators.HW_Fine AS Biomasa_podziemna,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag+
        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.SlowBG AS Martwa_materia_organiczna_DOM,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SlowAG+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag AS Nadziemna_martwa_mat_org_DOM,

        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.SlowBG AS Podziemna_martwa_mat_org_DOM,

        tblPoolIndicators.FastBG+ 
        tblPoolIndicators.Medium+ 
        tblPoolIndicators.SWStemSnag+ 
        tblPoolIndicators.SWBranchSnag+ 
        tblPoolIndicators.HWStemSnag+ 
        tblPoolIndicators.HWBranchSnag AS Martwe_drewno,

        tblPoolIndicators.VFastAG+ 
        tblPoolIndicators.FastAG+ 
        tblPoolIndicators.SlowAG AS Ścioła,

        tblPoolIndicators.VFastBG+ 
        tblPoolIndicators.SlowBG AS Węgiel_w_glebie,

        tblPoolIndicators.VFastAG AS Nadziemna_b_sz_r_m_m_o_DOM, 
        tblPoolIndicators.VFastBG AS Podziemna_b_sz_r_m_m_o_DOM, 
        tblPoolIndicators.FastAG AS Nadziemna_sz_r_m_m_o_DOM, 
        tblPoolIndicators.FastBG AS Podziemna_sz_r_m_m_o_DOM, 
        tblPoolIndicators.Medium AS Średnio_r_m_m_o_DOM, 
        tblPoolIndicators.SlowAG AS Nadziemna_wolno_r_m_m_o_DOM, 
        tblPoolIndicators.SlowBG AS Podziemna_wolno_r_m_m_o_DOM, 
        tblPoolIndicators.SWStemSnag AS Iglasty_posusz_pnie, 
        tblPoolIndicators.SWBranchSnag AS Iglasty_posusz_gałęzie, 
        tblPoolIndicators.HWStemSnag AS Liściasty_posusz_pnie, 
        tblPoolIndicators.HWBranchSnag AS Liściasty_posusz_gałęzie, 
        tblPoolIndicators.SW_Merch AS Grubizna_iglasta, 
        tblPoolIndicators.SW_Foliage AS Aparat_asymilacyjny_igl, 
        tblPoolIndicators.SW_Other AS Inne_iglaste, 
        tblPoolIndicators.SW_Coarse AS Grube_korzenie_iglaste, 
        tblPoolIndicators.SW_Fine AS Cienkie_korzenie_iglaste, 
        tblPoolIndicators.HW_Merch AS Grubizna_liściasta, 
        tblPoolIndicators.HW_Foliage AS Aparat_asymilacyjny_liśc, 
        tblPoolIndicators.HW_Other AS Inne_liściaste, 
        tblPoolIndicators.HW_Coarse AS Grube_korzenie_liściaste, 
        tblPoolIndicators.HW_Fine AS Cienkie_korzenie_liściaste

        FROM tblUserDefdClassSets INNER JOIN tblPoolIndicators 
        ON tblUserDefdClassSets.UserDefdClassSetID = tblPoolIndicators.UserDefdClassSetID
        WHERE (((tblUserDefdClassSets.Name)<>'SO,999999,99-99,REF,9999999999'))
        ORDER BY tblPoolIndicators.PoolIndID, tblPoolIndicators.TimeStep;'''


def Process1(filename, outfile):
    # global start_time
    start_time = time.time()

    # adres pyodbc.connect() na podstawie wygenerowanego DSN w Control Panel -> Administrative Tools -> Data Sources.
    # conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'
    #                       r'UID=admin;'
    #                       r'UserCommitSync=Yes;'
    #                       r'Threads=3;'
    #                       r'SafeTransactions=0;'
    #                       r'PageTimeout=5;'
    #                       r'MaxScanRows=8;'
    #                       r'MaxBufferSize=2048;'
    #                       r'FIL={MS Access};'
    #                       r'DriverId=25;'
    #                       r'DefaultDir=C:\PythonScripts;'
    #                       r'DBQ=C:\PythonScripts\project.mdb;')

    # driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};UID=admin;UserCommitSync=Yes;Threads=3;SafeTransactions=0;PageTimeout=5;MaxScanRows=8;MaxBufferSize=2048;FIL={MS Access};DriverId=25;DefaultDir=C:\PythonScripts;DBQ=C:\PythonScripts\project.mdb'

    in_path = 'DBQ=' + filename
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_time_step)

    # global head_row, fetch
    head_row = cursor.description
    fetch = cursor.fetchall()

    wb = Workbook()

    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row)):
        sheet1.cell(row=1, column=i + 1).value = head_row[i][0]

    # po kolumnach
    for j in range(0, len(head_row)):
        for i in range(0, len(fetch)):
            sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]

    sheet1.insert_cols(2)
    sheet1.cell(row=1, column=2).value = "Delta_Ekosystemu"
    for i in range(0, 50):
        sheet1.cell(row=i + 3, column=2).value = sheet1.cell(row=i + 3, column=3).value - sheet1.cell(row=i + 2,
                                                                                                      column=3).value

    # *************GENEROWANIE WYKRESÓW

    wb.create_sheet("Model_CBM_wykresy")

    ws = wb["Model_CBM_wykresy"]

    pos = []
    for i in range(1, 170, 16):
        pos.append(str("A" + str(i)))
        pos.append(str("K" + str(i)))
        pos.append(str("U" + str(i)))

    for i in range(len(pos)):
        c1 = LineChart()
        c1.title = str(sheet1.cell(row=1, column=i + 2).value)
        c1.style = 10
        c1.y_axis.title = 'węgiel w tonach'
        c1.x_axis.title = 'rok symulacji'

        data = Reference(sheet1, min_col=i + 2, min_row=1, max_col=i + 2, max_row=52)
        c1.add_data(data, titles_from_data=True)

        cats = Reference(sheet1, min_col=1, min_row=2, max_row=52)
        c1.set_categories(cats)

        # formatowanie serii danych
        s1 = c1.series[0]
        s1.graphicalProperties.line.solidFill = "00AAAA"
        s1.graphicalProperties.line.width = 20000  # width in EMUs

        # titles font
        font = Font(typeface='Verdana')
        size = 1000  # 14 point size
        cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
        pp = ParagraphProperties(defRPr=cp)
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        c1.x_axis.txPr = rtp  # Works!
        c1.y_axis.txPr = rtp  # Works!

        # X and Y axes titles
        c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
        c1.y_axis.title.tx.rich.p[0].pPr = pp

        # font
        font_test = Font(typeface='Calibri')
        cp = CharacterProperties(latin=font_test, sz=1000)
        c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

        c1.title.txPr = rtp

        c1.height = 8
        c1.width = 16
        ws.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 10%")
    root.update()


def Process2(filename, outfile):
    # KodNadl
    start_time = time.time()

    in_path = 'DBQ=' + filename
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_kod_nadl)

    head_row = cursor.description
    fetch = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row)):
        sheet1.cell(row=1, column=i + 1).value = head_row[i][0]

    # po kolumnach
    for j in range(0, len(head_row)):
        for i in range(0, len(fetch)):
            sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value
    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    for i in range(1, len(ls_uniq)):
        sheet_name = ls_uniq[i]
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 35
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == ls_uniq[i]:

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 20%")
    root.update()


def Process2_ha(filename, outfile):
    # KodNadl
    start_time = time.time()

    in_path = 'DBQ=' + filename
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_kod_nadl)

    head_row = cursor.description
    fetch = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row)):
        sheet1.cell(row=1, column=i + 1).value = head_row[i][0]

    # po kolumnach
    for j in range(0, len(head_row)):
        for i in range(0, len(fetch)):
            if j >= 2:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]/Nadl_pow[sheet1.cell(row=i+2, column=1).value]
            else:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]
            # print(Nadl_pow[sheet1.cell(row=i+2, column=1).value])

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value
    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    for i in range(1, len(ls_uniq)):
        sheet_name = ls_uniq[i]
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 35
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == ls_uniq[i]:

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column

                        # sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                        # if c > 1 and type(sheet1.cell(row=j, column=c + 1).value) != str and sheet1.cell(row=j, column=c + 1).value != None:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value / Nadl_pow[ls_uniq[i]]  # podział na sume powierzchni wydzieleń poszczegolnych nadlesnictw
                        # else:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value

                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value

                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    #usunięcie arkusza bo nie przeliczony na hektar
    # wb.remove(wb["Model_CBM"])
    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 20%")
    root.update()


def Process2_roznica(outfile):
    # KodNadl
    start_time = time.time()

    in_path = 'DBQ=' + filename1
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_kod_nadl)

    head_row1 = cursor.description
    fetch1 = cursor.fetchall()

    in_path = 'DBQ=' + filename2
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_kod_nadl)
    head_row2 = cursor.description
    fetch2 = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row1)):
        sheet1.cell(row=1, column=i + 1).value = head_row1[i][0]

    # po kolumnach
    for j in range(0, len(head_row1)):
        for i in range(0, len(fetch1)):
            if j > 1 and type(fetch1) != str and type(fetch1) != None:
                sheet1.cell(row=i + 2, column=j + 1).value = round(fetch1[i][j], 5) - round(fetch2[i][j],
                                                                                            5)  # DIFFERENCE
            else:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch1[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value
    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    for i in range(1, len(ls_uniq)):
        sheet_name = ls_uniq[i]
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 35
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == ls_uniq[i]:

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 20%")
    root.update()


def Process2_roznica_ha(outfile):
    # KodNadl
    start_time = time.time()

    in_path = 'DBQ=' + filename1
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_kod_nadl)

    head_row1 = cursor.description
    fetch1 = cursor.fetchall()

    in_path = 'DBQ=' + filename2
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_kod_nadl)

    head_row2 = cursor.description
    fetch2 = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row1)):
        sheet1.cell(row=1, column=i + 1).value = head_row1[i][0]

    # po kolumnach
    for j in range(0, len(head_row1)):
        for i in range(0, len(fetch1)):
            if j > 1 and type(fetch1) != str and type(fetch1) != None:
                sheet1.cell(row=i + 2, column=j + 1).value = (round(fetch1[i][j], 5) - round(fetch2[i][j],5)) / Nadl_pow[sheet1.cell(row=i + 2, column=1).value]  # DIFFERENCE
            else:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch1[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value
    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    for i in range(1, len(ls_uniq)):
        sheet_name = ls_uniq[i]
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 35
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == ls_uniq[i]:

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        # if c > 1 and type(sheet1.cell(row=j, column=c + 1).value) != str and sheet1.cell(row=j, column=c + 1).value != None:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value / Nadl_pow[ls_uniq[i]]
                        #     # sh_m.cell(row=r + 2, column=c + 1).value = round(sheet1.cell(row=j, column=c + 1).value / DzDod_pow[ls_uniq[i]], 5)
                        # else:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 20%")
    root.update()


def Process3(filename, outfile):
    # Dz_Dod
    start_time = time.time()

    in_path = 'DBQ=' + filename
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_dzial_dod)

    head_row = cursor.description
    fetch = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row)):
        sheet1.cell(row=1, column=i + 1).value = head_row[i][0]

    # po kolumnach
    for j in range(0, len(head_row)):
        for i in range(0, len(fetch)):
            sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value

    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    for i in range(1, len(ls_uniq)):
        sheet_name = ls_uniq[i]
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 35
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == ls_uniq[i]:

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 40%")
    root.update()


def Process3_ha(filename, outfile):
    # Dz_Dod
    start_time = time.time()

    # DzDod_pow = {'REF': 2692.02,
    #         'W-MDREW': 191.23,
    #         'W-ODOD': 1617.04,
    #         'W-OGLO': 41.68,
    #         'W-OGLP': 35.53,
    #         'W-OINN': 18.76,
    #         'W-ONAT': 1270.69,
    #         'W-OOTW': 239.96,
    #         'W-PODS': 8132.172,
    #         'W-PRZPL': 10.66,
    #         'W-RNIEZ': 128.25,
    #         'W-SZYBD': 2.68,
    #         'W-SZYBK': 188.28,
    #         'W-WREBN': 55.19,
    #         'W-ZALES': 37.51
    #         }

    in_path = 'DBQ=' + filename
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_dzial_dod)

    head_row = cursor.description
    fetch = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row)):
        sheet1.cell(row=1, column=i + 1).value = head_row[i][0]

    # po kolumnach
    for j in range(0, len(head_row)):
        for i in range(0, len(fetch)):
            if j >= 2:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]/DzDod_pow[sheet1.cell(row=i+2, column=1).value]
            else:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]
            # print(Nadl_pow[sheet1.cell(row=i+2, column=1).value])

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value

    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    for i in range(1, len(ls_uniq)):
        sheet_name = ls_uniq[i]
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 35
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == ls_uniq[i]:

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column

                        # if c > 1 and type(sheet1.cell(row=j, column=c + 1).value) != str and sheet1.cell(row=j, column=c + 1).value != None:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value / DzDod_pow[ls_uniq[i]]  # podział na sume powierzchni wydzieleń poszczegolnych dzialan dodatkowych
                        # else:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 40%")
    root.update()


def Process3_roznica(outfile):
    # Dz_Dod
    start_time = time.time()

    in_path = 'DBQ=' + filename1
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_dzial_dod)

    head_row1 = cursor.description
    fetch1 = cursor.fetchall()

    in_path = 'DBQ=' + filename2
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_dzial_dod)
    head_row2 = cursor.description
    fetch2 = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row1)):
        sheet1.cell(row=1, column=i + 1).value = head_row1[i][0]

    # po kolumnach
    for j in range(0, len(head_row1)):
        for i in range(0, len(fetch1)):
            if j > 1 and type(fetch1) != str and type(fetch1) != None:
                sheet1.cell(row=i + 2, column=j + 1).value = round(fetch1[i][j], 5) - round(fetch2[i][j],
                                                                                            5)  # DIFFERENCE
            else:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch1[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value

    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    for i in range(1, len(ls_uniq)):
        sheet_name = ls_uniq[i]
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 35
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == ls_uniq[i]:

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 40%")
    root.update()


def Process3_roznica_ha(outfile):
    # Dz_Dod
    start_time = time.time()

    # DzDod_pow = {'REF': 2692.02,
    #              'W-MDREW': 191.23,
    #              'W-ODOD': 1617.04,
    #              'W-OGLO': 41.68,
    #              'W-OGLP': 35.53,
    #              'W-OINN': 18.76,
    #              'W-ONAT': 1270.69,
    #              'W-OOTW': 239.96,
    #              'W-PODS': 8132.172,
    #              'W-PRZPL': 10.66,
    #              'W-RNIEZ': 128.25,
    #              'W-SZYBD': 2.68,
    #              'W-SZYBK': 188.28,
    #              'W-WREBN': 55.19,
    #              'W-ZALES': 37.51
    #              }

    in_path = 'DBQ=' + filename1
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_dzial_dod)

    head_row1 = cursor.description
    fetch1 = cursor.fetchall()

    in_path = 'DBQ=' + filename2
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_dzial_dod)

    head_row2 = cursor.description
    fetch2 = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row1)):
        sheet1.cell(row=1, column=i + 1).value = head_row1[i][0]

    # po kolumnach
    for j in range(0, len(head_row1)):
        for i in range(0, len(fetch1)):
            if j > 1 and type(fetch1) != str and type(fetch1) != None:
                sheet1.cell(row=i + 2, column=j + 1).value = (round(fetch1[i][j], 5) - round(fetch2[i][j],5)) / DzDod_pow[sheet1.cell(row=i + 2, column=1).value]  # DIFFERENCE
            else:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch1[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value

    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    for i in range(1, len(ls_uniq)):
        sheet_name = ls_uniq[i]
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 35
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == ls_uniq[i]:

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        # if c > 1 and type(sheet1.cell(row=j, column=c + 1).value) != str and sheet1.cell(row=j, column=c + 1).value != None:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value / DzDod_pow[ls_uniq[i]]
                        #     # sh_m.cell(row=r + 2, column=c + 1).value = round(sheet1.cell(row=j, column=c + 1).value / DzDod_pow[ls_uniq[i]], 5)
                        # else:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 40%")
    root.update()


def Process4(filename, outfile):
    # Arodes

    # 1008017155
    arod_graph = arodes_nr
    # arod_graph = input("(Opcjonalnie) Podaj numer wewnetrzny wydzielenia dla ktorego wygenerowac wykres : ")

    start_time = time.time()

    in_path = 'DBQ=' + filename
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_arodes)

    head_row = cursor.description
    fetch = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row)):
        sheet1.cell(row=1, column=i + 1).value = head_row[i][0]

    # po kolumnach
    for j in range(0, len(head_row)):
        for i in range(0, len(fetch)):
            sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value

    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    # if arod_graph.isdigit():

    for i in range(0, len(ls_modified_arodes)):
        # if arod_graph == ls_uniq[i]:
        sheet_name = str(ls_modified_arodes[i])
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 34
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715

            if sheet1.cell(row=j, column=1).value == str(ls_modified_arodes[i]):

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 80%")
    root.update()


def Process4_ha(filename, outfile):
    # Arodes

    # 1008017155
    arod_graph = arodes_nr
    # arod_graph = input("(Opcjonalnie) Podaj numer wewnetrzny wydzielenia dla ktorego wygenerowac wykres : ")

    start_time = time.time()

    in_path = 'DBQ=' + filename
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_arodes)

    head_row = cursor.description
    fetch = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row)):
        sheet1.cell(row=1, column=i + 1).value = head_row[i][0]

    # po kolumnach
    for j in range(0, len(head_row)):
        for i in range(0, len(fetch)):
            if j >= 2:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]/Arodes_pow[int(sheet1.cell(row=i+2, column=1).value)]
            else:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value

    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    # if arod_graph.isdigit():

    for i in range(0, len(ls_modified_arodes)):
        # if arod_graph == ls_uniq[i]:
        sheet_name = str(ls_modified_arodes[i])
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 34
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == str(ls_modified_arodes[i]):

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        # if c > 1 and type(sheet1.cell(row=j, column=c + 1).value) != str and sheet1.cell(row=j, column=c + 1).value != None:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value / Arodes_pow[ls_modified_arodes[i]]  # podział na sume powierzchni wydzieleń arodes
                        # else:
                        #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 80%")
    root.update()


def Process4_roznica(outfile):
    # Arodes

    # 1008017155
    arod_graph = arodes_nr
    # arod_graph = input("(Opcjonalnie) Podaj numer wewnetrzny wydzielenia dla ktorego wygenerowac wykres : ")

    start_time = time.time()

    in_path = 'DBQ=' + filename1
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_arodes)

    head_row1 = cursor.description
    fetch1 = cursor.fetchall()

    in_path = 'DBQ=' + filename2
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_arodes)

    head_row2 = cursor.description
    fetch2 = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row1)):
        sheet1.cell(row=1, column=i + 1).value = head_row1[i][0]

    # po kolumnach
    for j in range(0, len(head_row1)):
        for i in range(0, len(fetch1)):
            if j > 1 and type(fetch1) != str and type(fetch1) != None:
                sheet1.cell(row=i + 2, column=j + 1).value = round(fetch1[i][j], 5) - round(fetch2[i][j],
                                                                                            5)  # DIFFERENCE
            else:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch1[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value

    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    # if arod_graph.isdigit():

    for i in range(0, len(ls_modified_arodes)):
        # if arod_graph == ls_uniq[i]:
        sheet_name = str(ls_modified_arodes[i])
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 34
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == str(ls_modified_arodes[i]):

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 80%")
    root.update()


def Process4_roznica_ha(outfile):
    # Arodes

    # 1008017155
    arod_graph = arodes_nr
    # arod_graph = input("(Opcjonalnie) Podaj numer wewnetrzny wydzielenia dla ktorego wygenerowac wykres : ")

    start_time = time.time()

    in_path = 'DBQ=' + filename1
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_arodes)

    head_row1 = cursor.description
    fetch1 = cursor.fetchall()

    in_path = 'DBQ=' + filename2
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_agreg_arodes)

    head_row2 = cursor.description
    fetch2 = cursor.fetchall()

    wb = Workbook()
    wb.create_sheet("Model_CBM")
    sheet1 = wb["Model_CBM"]
    wb.remove(wb["Sheet"])

    # WPISANIE nagłowków
    for i in range(0, len(head_row1)):
        sheet1.cell(row=1, column=i + 1).value = head_row1[i][0]

    # po kolumnach
    for j in range(0, len(head_row1)):
        for i in range(0, len(fetch1)):
            if j > 1 and type(fetch1) != str and type(fetch1) != None:
                sheet1.cell(row=i + 2, column=j + 1).value = (round(fetch1[i][j], 5) - round(fetch2[i][j],5)) / Arodes_pow[int(sheet1.cell(row=i + 2, column=1).value)]  # DIFFERENCE
            else:
                sheet1.cell(row=i + 2, column=j + 1).value = fetch1[i][j]

    sheet1.insert_cols(3)
    sheet1.cell(row=1, column=3).value = "Delta_Ekosystemu"
    for i in range(0, sheet1.max_row - 2):
        if sheet1.cell(row=i + 3, column=2).value != 0:
            sheet1.cell(row=i + 3, column=3).value = sheet1.cell(row=i + 3, column=4).value - sheet1.cell(row=i + 2,
                                                                                                          column=4).value

    ls = []
    for cell in sheet1['A']:
        ls.append(cell.value)

    ls_uniq = []
    for i in range(len(ls)):
        if ls[i] not in ls_uniq:
            ls_uniq.append(ls[i])

    # if arod_graph.isdigit():

    for i in range(0, len(ls_modified_arodes)):
        # if arod_graph == ls_uniq[i]:
        sheet_name = str(ls_modified_arodes[i])
        wb.create_sheet(sheet_name)
        sh_m = wb[sheet_name]

        for h in range(sheet1.max_column):  # 34
            sh_m.cell(row=1, column=h + 1).value = sheet1.cell(row=1, column=h + 1).value

        first_row = 0
        for j in range(1, sheet1.max_row + 1):  # 715
            if sheet1.cell(row=j, column=1).value == str(ls_modified_arodes[i]):

                for r in range(first_row, 51):  # 51
                    for c in range(35):  # 35 sheet1.max_column
                        for c in range(35):  # 35 sheet1.max_column
                            # if c > 1 and type(sheet1.cell(row=j, column=c + 1).value) != str and sheet1.cell(row=j, column=c + 1).value != None:
                            #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value / Arodes_pow[ls_modified_arodes[i]]  # podział na sume powierzchni wybranych wydzieleń arodes
                            # else:
                            #     sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                            sh_m.cell(row=r + 2, column=c + 1).value = sheet1.cell(row=j, column=c + 1).value
                first_row += 1

        pos = []
        for i in range(1, 170, 16):
            pos.append(str("AJ" + str(i)))
            pos.append(str("AT" + str(i)))
            pos.append(str("BD" + str(i)))

        for i in range(len(pos)):
            c1 = LineChart()
            c1.title = sheet_name + " " + str(sh_m.cell(row=1, column=i + 3).value)
            c1.style = 10
            c1.y_axis.title = 'węgiel w tonach'
            c1.x_axis.title = 'rok symulacji'

            data = Reference(sh_m, min_col=i + 3, min_row=1, max_col=i + 3, max_row=52)
            c1.add_data(data, titles_from_data=True)

            cats = Reference(sheet1, min_col=2, min_row=2, max_row=52)
            c1.set_categories(cats)

            # formatowanie serii danych
            s1 = c1.series[0]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.width = 20000  # width in EMUs

            # titles font
            font = Font(typeface='Verdana')
            size = 1000  # 14 point size
            cp = CharacterProperties(latin=font, sz=size, b=False)  # Not bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            c1.x_axis.txPr = rtp  # Works!
            c1.y_axis.txPr = rtp  # Works!

            # X and Y axes titles
            c1.x_axis.title.tx.rich.p[0].pPr = pp  # Works!
            c1.y_axis.title.tx.rich.p[0].pPr = pp

            # font
            font_test = Font(typeface='Calibri')
            cp = CharacterProperties(latin=font_test, sz=1000)
            c1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            c1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

            c1.title.txPr = rtp

            c1.height = 8
            c1.width = 16
            sh_m.add_chart(c1, pos[i])

    out_path = folder_selected + "/"
    wb.save(out_path + outfile)

    elapsed_time1 = time.time() - start_time
    print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
    print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

    proces.configure(text="Processing... 80%")
    root.update()


def Process5(filename, outfile):
    # tblPoolIndicators
    start_time = time.time()

    in_path = 'DBQ=' + filename
    driver_param = 'Driver={Microsoft Access Driver (*.mdb, *.accdb)};%s' % (in_path)
    conn = pyodbc.connect(driver_param)
    cursor = conn.cursor()

    cursor.execute(SQL_tblPoolIndicators)

    head_row = cursor.description
    fetch = cursor.fetchall()

    mem = psutil.virtual_memory()
    MEMORY_LIMIT = 20 * 1024 * 1024 * 1024  # 20GB

    if mem.available > MEMORY_LIMIT:
        print("20 GB memory available. Proceeding XLSX generation...")
        try:
            wb = Workbook()
            wb.create_sheet("Model_CBM")
            sheet1 = wb["Model_CBM"]
            wb.remove(wb["Sheet"])

            # WPISANIE nagłowków
            for i in range(0, len(head_row)):
                sheet1.cell(row=1, column=i + 1).value = head_row[i][0]

            # WPISANIE danych

            # po wierszach
            # for i in range(0,len(fetch)):
            #     for j in range(0,len(head_row)):
            #         sheet1.cell(row=i+2,column=j+1).value = fetch[i][j]

            # po kolumnach
            for j in range(0, len(head_row)):
                for i in range(0, len(fetch)):
                    sheet1.cell(row=i + 2, column=j + 1).value = fetch[i][j]
                    # if mem.available <= MEMORY_LIMIT:
                    #     del sheet1
                    #     Proces2csv()
                    #     return

            out_path = folder_selected + "/"
            wb.save(out_path + outfile)

            elapsed_time1 = time.time() - start_time
            print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
            print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))
            proces.configure(text="Processing... 100%")
            root.update()

        except Exception as e:
            print(e)
    else:
        print("20 GB memory NOT available. Proceeding CSV generation and conversion to XLSX...")

        # ************* zamiana znaków z Ś na S (opcjonalne)
        # zamiana tupli w liście na listy w liście
        fetch_list_lists = []
        for i in range(len(fetch)):
            sub_list = []
            for j in range(len(fetch[i])):
                sub_list.append(fetch[i][j])
            fetch_list_lists.append(sub_list)

        # zamiana znaków
        for i in range(0, len(fetch_list_lists)):
            fetch_list_lists[i][0] = str(fetch_list_lists[i][0]).replace("Ś", "S")
            fetch_list_lists[i][1] = str(fetch_list_lists[i][1]).replace("Ś", "S")

        # zamiana list w liście na tuple w liście
        fetch_list_tuples = []
        for i in range(len(fetch_list_lists)):
            fetch_list_tuples.append(tuple(fetch_list_lists[i]))
        # *************** koniec zamiany znaków

        # rozwiązanie sprawy z nagłówkami
        list_head_rows = []
        for i in range(len(head_row)):
            list_head_rows.append(head_row[i][0])

        list_head_rows = tuple(list_head_rows)
        ls = []
        ls.append(list_head_rows)

        # zapisanie pliku csv
        try:
            csv_path = folder_selected + "/" + outfile[:-5] + ".csv"
            with open(csv_path, 'w', newline='') as csv_file:
                csv_writer = csv.writer(csv_file)  # default field-delimiter is ","
                csv_writer.writerows(ls)
                csv_writer.writerows(fetch_list_tuples)
        except Exception as e:
            print(e)

        # elapsed1
        elapsed_time1 = time.time() - start_time
        print("Export to CSV successful. Elapsed time: ")
        print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))

        # konwersja istniejącego pliku CSV na excel
        try:
            xlsx_path = folder_selected + "/" + outfile
            df_new = pd.read_csv(csv_path, encoding='utf-8')
            writer = pd.ExcelWriter(xlsx_path, engine='xlsxwriter')
            df_new.to_excel(writer, index=False, encoding='utf-8')
            writer.save()

            elapsed_time1 = time.time() - start_time
            print(f"Export to {outfile} successful. Time: {datetime.datetime.now()}. Elapsed time: ")
            print(time.strftime("%H:%M:%S", time.gmtime(elapsed_time1)))
            proces.configure(text="Processing... 100%")
            root.update()

        except Exception as e:
            print(e)


def Process6():
    mem()
    # del head_row, fetch
    print("All files exported at:")
    print(datetime.datetime.now())
    proces.configure(text="Processing... 100%")
    root.update()

    # zakmnięcie okna, zwolnienie zajętej pamięci
    root.destroy()
    # del root
    mem()


def SelectFile():
    global filename1
    filename1 = filedialog.askopenfilename(initialdir=r"C:\Program Files (x86)\Operational-Scale CBM-CFS3\Projects",
                                           title="Select project result file",
                                           filetypes=((".mdb files", "*.mdb"), ("all files", "*.*")))
    print("Projekt dz dod: " + filename1)


def SelectFile2():
    global filename2
    filename2 = filedialog.askopenfilename(initialdir=r"C:\Program Files (x86)\Operational-Scale CBM-CFS3\Projects",
                                           title="Select second project result file",
                                           filetypes=((".mdb files", "*.mdb"), ("all files", "*.*")))
    print("Projekt bez dod: " + filename2)


def SelectInputFile():
    global input_filename
    input_filename = filedialog.askopenfilename(initialdir=r"/", title="Select project result file",
                                                filetypes=((".xls files", "*.xls"), ("all files", "*.*")))
    print("Plik wsadowy: " + input_filename)


def SelectOutputFolder():
    global folder_selected
    folder_selected = filedialog.askdirectory(title="Select output folder")
    print(f"Wyniki: {folder_selected}")


def submitArodes():
    global arodes_nr, ls_modified_arodes
    # arodes.configure(text="arodes number(s) submitted.")
    # button4.configure(state=DISABLED)
    # arodes_nr = entry_1.get()

    arodes_nr = entry_1.get()
    # print(arodes_nr)

    arodes_list_str = arodes_nr.split(',')

    # print(arodes_list_str)

    arodes_list_int = []
    for el in arodes_list_str:
        arodes_list_int.append(int(el))

    ls_modified_arodes = arodes_list_int

    print(arodes_list_int)


GUI()
